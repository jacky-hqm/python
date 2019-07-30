
# -*- coding: utf-8 -*-
import re
from bs4 import BeautifulSoup
import requests
import time

from openpyxl.workbook import Workbook

outwb = Workbook()
careerSheet = outwb.create_sheet('career',0)

url1='https://myip.ms/browse/sites/'
url2='/ipID/23.227.38.0/ipIDii/23.227.38.255/sort/2/asc/1/'

PAGE_list=[]
ret11=[]
su11=[]
ret1=[]

for x in range(1,100):
    url=url1+str(x)+url2

    PAGE_list.append(url)

def datas(url):

    headers={'Referer': 'https://myip.ms/browse/sites/1/ipID/23.227.38.0/ipIDii/23.227.38.255/sort/2/asc/1/',
             'Host': 'myip.ms',
             'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 UBrowser/6.2.3964.2 Safari/537.36',
             'Upgrade-Insecure-Requests': '1',
             'Cookie': 's2_uLang=en; PHPSESSID=5mqgji0fp7ofjpdb3iff03l0n6; s2_uID=46963; s2_uKey=9b9805dcd48d3c91bebc173b834a1c2bcd28c0ef; s2_uDer=41922ff8aa290c1e356b670aff3ba2d12ac81fa6; s2_theme_ui=red; s2_csrf_cookie_name=87a6294595fd2e2e7e076729c233768f; __unam=737437c-164dab1df8d-4b9f36b8-76; sw=190.3; sh=32.9; s2_csrf_cookie_name=87a6294595fd2e2e7e076729c233768f; _ga=GA1.2.1568725569.1532677447; _gid=GA1.2.402395902.1532677447',
             'X-Powered-By':'PleskLin',
             }
    req=requests.get(url,headers=headers)
    soup=BeautifulSoup(req.content,'lxml')
    biao=soup.find_all('td',class_='row_name')
    t=reduce(lambda x, y:str(x)+str(y),biao)
    patt1='[^/]*>[^>]+</a'
    ret=re.findall(patt1,t)
    t1=reduce(lambda x, y:str(x)+str(y),ret)
    ret1=re.findall(r'[>](.*?)[<]', t1)

    sulian=soup.find_all('span',attrs={'class':'bold arial grey'})
    su=reduce(lambda x, y:str(x)+str(y),sulian)
    su1=re.findall(r'[#](.*?)[<]', su)

    ret11.extend(ret1)
    su11.extend(su1)
    print 1

def main():
    for url in PAGE_list:
        datas(url)
if __name__=='__main__':
    main()

datas(url)
b=0
c=0
print ret11

for i in ret11 :
   b=b+1
   careerSheet.cell(row=b,column=1).value =i
   outwb.save("sample.xlsx")
for j in su11 :
   c=c+1
   careerSheet.cell(row=c,column=2).value =j
   outwb.save("sample.xlsx")
